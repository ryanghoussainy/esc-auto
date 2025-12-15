import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from copy import copy

# flag to control sheet protection copying
PROTECT_SHEET = False

def amindefy_timesheets(timesheet_folder: str, output_file: str, progress_callback, error_callback):
    """
    Combines Excel files into one workbook while preserving ALL formatting,
    including protected and hidden columns.
    """
    try:
        # Create a new workbook
        output_wb = Workbook()
        output_wb.remove(output_wb.active)  # Remove default sheet
        
        for filename in os.listdir(timesheet_folder):
            # Skip non-Excel files
            if not filename.endswith(".xlsx"):
                continue
            
            amindefy_timesheet(filename, timesheet_folder, output_wb)

            progress_callback(f"Added timesheet: {filename}\n")
        
        # Save the output workbook
        output_wb.save(output_file)
        output_wb.close()
        
        progress_callback(f"\n✅ TIMESHEETS COMBINED SUCCESSFULLY! Output file: {output_file}\n")
    
    except Exception as e:
        error_callback(f"❌ ERROR: {str(e)}", "red")


def amindefy_timesheet(filename: str, timesheet_folder: str, output_wb: Workbook):
    """
    Adds a single timesheet to an existing workbook while preserving ALL formatting.
    """
    file_path = os.path.join(timesheet_folder, filename)
    
    # Load source workbook
    source_wb = load_workbook(file_path)
    source_ws = source_wb.active
    
    # Create sheet name from filename
    sheet_name = os.path.splitext(filename)[0]

    # remove dashes from sheet name
    sheet_name = sheet_name.replace('-', ' ')

    # get a sheet prefix for named ranges
    sheet_prefix = sheet_name.replace(' ', '_')
    
    # Use copy_worksheet to get most formatting automatically
    output_ws = output_wb.create_sheet(title=sheet_name)

    # Copy all cell values and styles
    for row in source_ws.iter_rows():
        for cell in row:
            # Replace RATE and GALA by sheet-specific named ranges
            if isinstance(cell.value, str) and 'RATE' in cell.value and 'VLOOKUP' in cell.value:
                updated_value = cell.value.replace('RATE', f'{sheet_prefix}_RATE')
                new_cell = output_ws.cell(row=cell.row, column=cell.column, value=updated_value)
            elif isinstance(cell.value, str) and 'GALA' in cell.value and 'VLOOKUP' in cell.value:
                updated_value = cell.value.replace('GALA', f'{sheet_prefix}_GALA')
                new_cell = output_ws.cell(row=cell.row, column=cell.column, value=updated_value)
            else:
                new_cell = output_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Copy all cell styling
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)

                # Copy fill but change theme colour
                if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    if cell.fill.fgColor.type == 'theme' and cell.fill.fgColor.theme == 9:
                        # Change theme colour
                        new_cell.fill = PatternFill(start_color="DBF2D0", end_color="DBF2D0", fill_type="solid")
                    else:
                        # Copy the existing fill if it's not a theme colour
                        new_cell.fill = copy(cell.fill)

                new_cell.number_format = copy(cell.number_format)
                if PROTECT_SHEET:
                    new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            
            # Copy hyperlinks if present
            if cell.hyperlink:
                new_cell.hyperlink = copy(cell.hyperlink)
            
            # Copy comments if present
            if cell.comment:
                new_cell.comment = copy(cell.comment)
    
    # Copy merged cells
    for merged_cell_range in source_ws.merged_cells.ranges:
        output_ws.merge_cells(str(merged_cell_range))
    
    # Copy column dimensions (including hidden and width)
    for col_letter, col_dimension in source_ws.column_dimensions.items():
        output_ws.column_dimensions[col_letter].width = col_dimension.width
        output_ws.column_dimensions[col_letter].hidden = col_dimension.hidden
        output_ws.column_dimensions[col_letter].bestFit = col_dimension.bestFit
        output_ws.column_dimensions[col_letter].auto_size = col_dimension.auto_size
    
    # Hardcode column width for dropdown columns
    output_ws.column_dimensions['D'].width = 12.67
    output_ws.column_dimensions['F'].width = 15.5
    
    # Copy row dimensions (including hidden and height)
    for row_num, row_dimension in source_ws.row_dimensions.items():
        output_ws.row_dimensions[row_num].height = row_dimension.height
        output_ws.row_dimensions[row_num].hidden = row_dimension.hidden
    
    # Copy sheet properties
    output_ws.sheet_format = copy(source_ws.sheet_format)
    output_ws.sheet_properties = copy(source_ws.sheet_properties)
    
    # Copy page setup and print settings
    output_ws.page_setup = copy(source_ws.page_setup)
    output_ws.page_margins = copy(source_ws.page_margins)
    output_ws.print_options = copy(source_ws.print_options)
    
    # Copy sheet protection (if protected)
    if PROTECT_SHEET and source_ws.protection.sheet:
        output_ws.protection = copy(source_ws.protection)
    
    # Copy freeze panes
    if source_ws.freeze_panes:
        output_ws.freeze_panes = source_ws.freeze_panes
    
    # Copy sheet views (zoom, selection, etc.)
    if source_ws.views and source_ws.views.sheetView:
        output_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
        output_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal
        output_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    
    # Copy data validations
    if source_ws.data_validations:
        for dv in source_ws.data_validations.dataValidation:
            output_ws.data_validations.append(copy(dv))
    
    # Copy conditional formatting
    if source_ws.conditional_formatting:
        for range_string, rules in source_ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                output_ws.conditional_formatting.add(range_string, copy(rule))
    
    # Copy filters
    if source_ws.auto_filter:
        output_ws.auto_filter.ref = source_ws.auto_filter.ref
    
    # Copy tables
    for table in source_ws.tables.values():
        output_ws.add_table(copy(table))
    
    # Copy named ranges for GALA and RATE tables
    for name, named_range in source_wb.defined_names.items():
        # Ignore all non GALA and RATE named ranges
        if not (('GALA' in name) or ('RATE' in name)):
            continue

        # Create a new named range in the output workbook
        new_name = copy(named_range)

        # Update the named ranges to be unique per sheet
        new_name.name = f"{sheet_prefix}_{name}"

        # Update sheet reference to point to the new sheet name
        if named_range.value:
            # Replace old sheet name with new sheet name in the reference
            old_sheet_name = source_ws.title

            # sheet name must be quoted if it contains spaces
            if ' ' in old_sheet_name:
                old_sheet_name = f"'{old_sheet_name}'"
            if ' ' in sheet_name:
                quoted_sheet_name = f"'{sheet_name}'"
            
            new_name.value = named_range.value.replace(old_sheet_name, quoted_sheet_name)

        # Make the named range local to the new sheet
        new_name.localSheetId = output_wb.sheetnames.index(sheet_name)
        
        # Add to output workbook using dictionary assignment
        output_wb.defined_names[new_name.name] = new_name
    
    source_wb.close()
