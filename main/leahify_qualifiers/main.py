'''
This file's purpose is to turn Sammy's version of qualifiers into Leah's version,
hence the name leahify.
First, read leah's version, then for each swimmer in that sheet, find the corresponding
time in Sammy's version and add it to the dictionary.
'''

from .extract_tables import extract_tables, concat_tables
from reusables import match_swimmer, parse_name, get_event_name, is_final, rename_final_column
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re

TIME_COLUMN_INDEX = 5 # Index of the time column in Leah's tables

GROUPS = ["Dolphins", "Herrings", "Seals", "Piranhas", "Marlins"]

QUAL_TABLE_ID = "First name"  # Identifier for the qualifiers table
LEAH_TABLE_ID = "Lane"        # Identifier for the normal tables

REGEX_AGE_RANGE_LEAH = r"\b\d{1,2}\s*&\s*(Under|Over|under|over)|\b\d{1,2}\s*-\s*\d{1,2}" # Regex for age range


def get_qualifiers_table(file: str, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    '''
    Extract the tables from Sammy's version of the qualifiers.
    '''
    qualifiers_table, _, s_info = extract_tables(file, sheet_name, [(QUAL_TABLE_ID, 0)])
    return concat_tables(qualifiers_table), s_info

def get_leah_tables(file: str, sheet_name: str) -> tuple[list[pd.DataFrame], list[str], dict[str, (int, int, str)]]:
    '''
    Extract the tables from Leah's version of the qualifiers.
    
    Note: the length of the events will be equal to the length
    of the normal tables, so does not include the extra tables.

    Returns a tuple of:
    - A list of Tables (DataFrames) for each event
    - A list of full event names
    - A list of swimmer info (age from, age to, gender)
    '''
    return extract_tables(file, sheet_name, [(LEAH_TABLE_ID, 0)], get_events=True)

def load_qualifiers(sfile: str) -> tuple[pd.DataFrame, dict]:
    '''
    Load all qualifiers tables and swimmer info (age from, age to, gender) from the Excel file.
    '''
    qualifiers_tables = []
    swimmer_info = {}
    for group in GROUPS:
        qualifiers_table, s_info = get_qualifiers_table(sfile, group)
        qualifiers_tables.append(qualifiers_table)
        swimmer_info.update(s_info)
    # Concatenate all tables into a single table
    qualifiers_tables = pd.concat(qualifiers_tables, ignore_index=True)

    return qualifiers_tables, swimmer_info

def match_swimmers(
    qualifiers_table: pd.DataFrame,
    leah_tables: list[pd.DataFrame],
    events: list[str],
    time_column_name: str,
    confirm_callback,
    progress_callback,
) -> dict:
    '''
    Match swimmers from Leah's version to Sammy's version.
    Returns a dictionary mapping swimmer names (Sammy's version) to a list of events they swam.
    '''
    # For each swimmer in Leah's version, find the corresponding time in Sammy's version
    # Keep track of number of successful matches and number of ignored swimmers
    total, num_matches, num_ignored = 0, 0, 0

    # Keep track of manual and automatic matches
    manual_matches = {} # manual matches are from user input
    automatic_matches = {} # automatic matches are from exact matches
    ignored_swimmers = set() # keep track of ignored swimmers

    # Keep track of which events have been matched.
    # This is for the extras table (i.e. which swimmers swam but did not sign up)
    matched_events = {}  # Map from swimmer name (sammy's version) to a list of events

    # Each table corresponds to an event
    # Iterate over each table and each row to get each swimmer
    for tableIdx in range(len(leah_tables)):
        # Get the event
        event = events[tableIdx]
    
        ltable = leah_tables[tableIdx]

        progress_callback(f"Processing event: {event}", "yellow")

        # Iterate over each row in the table (i.e. each swimmer)
        for lrowIdx, lrow in ltable.iterrows():
            # Skip NAN rows
            if pd.isnull(lrow["Name"]):
                continue

            # Skip Northolt and St Helens swimmers
            if str(lrow["Team"]).strip() in ["Northolt", "St Helens"]:
                continue

            # Increment total number of swimmer rows
            total += 1

            # Parse the swimmer's name (Leah's version)
            lfirst_names, lsurname = parse_name(lrow['Name'])
            lfirst_name = lfirst_names.split()[0]

            # check if swimmer has already been ignored
            if (lfirst_name, lsurname) in ignored_swimmers:
                continue

            # Match the swimmer to Sammy's version
            swimmer = match_swimmer(
                lfirst_name,
                lsurname,
                qualifiers_table,
                automatic_matches,
                manual_matches,
                progress_callback=progress_callback,
                confirm_callback=confirm_callback
            )
            # Check if swimmer was ignored
            if swimmer.empty:
                ignored_swimmers.add((lfirst_name, lsurname))
                num_ignored += 1
                continue
            
            # Increment number of matches
            num_matches += 1

            # Add the event to the matched events
            sfirst_name = swimmer["First name"].values[0]
            ssurname = swimmer["Surname"].values[0]
            key = (sfirst_name, ssurname)
            matched_events.setdefault(key, []).append(event)

            progress_callback(f"Successfully matched: {lfirst_name.capitalize()} {lsurname.capitalize()}", "green")

            # Get the time
            time = swimmer[event].values[0]

            # Set the column type to string
            leah_tables[tableIdx] = leah_tables[tableIdx].astype({time_column_name: str})
            
            # Set the time if it's not nan
            leah_tables[tableIdx].at[lrowIdx, time_column_name] = time if not pd.isnull(time) else "DNS"

    progress_callback(f"Number of matches: {num_matches}/{total}")
    progress_callback(f"Number of ignored swimmers: {num_ignored}/{total}")

    return matched_events

def combine_tables(leah_tables: list[pd.DataFrame], time_column_name: str) -> pd.DataFrame:
    '''
    Concatenate all tables into a single table.
    '''
    ltable = leah_tables[0]
    output_table = pd.DataFrame(columns=ltable.columns)

    header_row = pd.DataFrame([ltable.columns], columns=ltable.columns)

    for tableIdx in range(len(leah_tables)):
        # Add the header row
        ltable = leah_tables[tableIdx]
        ltable_with_header = pd.concat([ltable.iloc[:1], header_row, ltable.iloc[1:]])
        output_table = pd.concat([output_table, ltable_with_header])

    # Get rid of nan values in the Time column
    output_table[time_column_name] = output_table[time_column_name].replace("nan", "")

    return output_table

def get_extras_per_event(
        qualifiers_table: pd.DataFrame,
        events: list[str],
        swimmer_info: dict,
        matched_events: dict,        
) -> dict:
    # Keep track of extra swimmers for each event
    extras_per_event = {} # Map from (event name, age range, gender) to a list of swimmers

    for idx, srow in qualifiers_table.iterrows():
        # Get all events swam by the swimmer
        events_swam = set(event for event in set(events) if not pd.isnull(srow[event]) and srow[event] != "DNS")

        # Missing events
        missing_events = []

        # Check if they have been matched for the same events
        for ev in events_swam:
            if (srow["First name"], srow["Surname"]) not in matched_events or ev not in matched_events[(srow["First name"], srow["Surname"])]:
                missing_events.append(ev)

        # Add the swimmer to the extras per event
        for missing_event in missing_events:
            # Get swimmer age range and gender
            age_from, age_to, gender = swimmer_info[(srow["First name"], srow["Surname"])]

            # If it's a final event, we make the age range 0-99
            if is_final(missing_event):
                age_from = 0
                age_to = 99
            
            # Add the swimmer to the extras per event table
            extras_per_event.setdefault((missing_event, age_from, age_to, gender), []).append(srow)

    return extras_per_event

def add_extras_to_leah_tables(
    leah_tables: list[pd.DataFrame],
    extras_per_event: dict,
) -> list[pd.DataFrame]:
    """
    Insert extras into each Leah table before combining.
    """
    
    for i in range(len(leah_tables)):
        leah_table = leah_tables[i]
        event_cell = leah_table.iloc[0, 0]

        # Extract event name
        event_name = get_event_name(event_cell)

        # Extract age range
        age_range = re.search(REGEX_AGE_RANGE_LEAH, event_cell)
        if age_range:
            age_range = age_range.group(0)
            if "under" in age_range.lower():
                age_from = 0
                age_to = int(age_range.split("&")[0].strip())
            elif "over" in age_range.lower():
                age_from = int(age_range.split("&")[0].strip())
                age_to = 99
            else:
                if "-" in age_range:
                    age_from, age_to = map(int, age_range.split("-"))
                elif "/" in age_range:
                    age_from, age_to = map(int, age_range.split("/"))
                else:
                    raise ValueError(f"Could not extract age range. Did not find - or /: {event_cell}")
        elif is_final(event_cell):
            age_from = 0
            age_to = 99
        else:
            raise ValueError(f"Could not extract age range from event name: {event_cell}")

        gender = "boys" if "boys" in event_cell.lower() else "girls"
        key = (event_name, age_from, age_to, gender)

        # If extras_per_event has an entry, add extras to the table
        if key in extras_per_event:
            # Add extra label row
            extra_label_row = pd.DataFrame([["EXTRA"] + [""] * (len(leah_table.columns) - 1)], columns=leah_table.columns)
            leah_tables[i] = pd.concat([leah_tables[i], extra_label_row], ignore_index=True)

            # For each extra swimmer, create a row
            for extra_row in extras_per_event[key]:
                # Add the correct columns to insert into Leah table
                extra_row = extra_row[["First name", "Surname", "ASA", "DOB", "Group", event_name]]
                
                # Pad the row with empty strings to match the number of columns in leah_table
                padded_row = list(extra_row.values) + [""] * (len(leah_table.columns) - len(extra_row.values))
                extra_row_df = pd.DataFrame([padded_row], columns=leah_table.columns)

                # Add it to the Leah table
                leah_tables[i] = pd.concat([leah_tables[i], extra_row_df], ignore_index=True)

    return leah_tables

def save_output_table_to_excel(
        output_table: pd.DataFrame,
        filename: str,
        time_column_name: str,
) -> None:
    '''
    Save the output table to an Excel file.
    '''
    # Add a border to the table
    wb = Workbook()
    ws = wb.active
    big_font = Font(size=18)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r in dataframe_to_rows(output_table, index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=1, max_row=len(output_table), min_col=1, max_col=len(output_table.columns)):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if cell.column == 6 and cell.value not in (time_column_name, "Finals"):
                cell.font = big_font
            if cell.row > 1 and cell.column == 1 and cell.value == "EXTRA":
                cell.fill = yellow_fill

    # Save the output table
    wb.save(filename)

def restore_final_column(output_table: pd.DataFrame):
    for idx in range(len(output_table)):
        row = output_table.iloc[idx]
        cell = row.iloc[0]
        # Skip non event header rows
        if not cell.startswith("Event"):
            continue

        # Skip non finals events
        if not is_final(get_event_name(cell)):
            continue
        
        # Change the time column name to "Finals"
        output_table.iloc[idx + 1, TIME_COLUMN_INDEX] = "Finals"

def add_time_column(leah_tables: list[pd.DataFrame]):
    for leah_table in leah_tables:
        if leah_table.columns[TIME_COLUMN_INDEX] not in ["Time", "Finals"]:
            leah_table.insert(TIME_COLUMN_INDEX, "Time", "")

def leahify_qualifiers(
    sfile: str,
    lfile: str,
    progress_callback,
    confirm_callback,
    error_callback,
    output_path: str = "output.xlsx",
) -> None:
    '''
    Turn Sammy's version of qualifiers into Leah's version.
    
    Args:
        sfile: Path to Sammy's qualifiers file
        lfile: Path to Leah's template file  
        output_path: Output file path
        progress_callback: Called with progress messages (str)
        confirm_callback: Called for user confirmations, expects (message: str, data: dict) -> str
        error_callback: Called with error messages (str)
    '''
    try:
        progress_callback("Loading qualifier times from Sammy's file...")
        
        # Load qualifier times (Sammy's version)
        qualifiers_table, swimmer_info = load_qualifiers(sfile)
        
        progress_callback("Extracting tables from Leah's template...")

        # Extract tables from Leah's version
        leah_tables, full_events, _ = get_leah_tables(lfile, None)
        
        # Add Time column if it doesn't exist
        add_time_column(leah_tables)

        # Get time column name in Leah's table
        time_column_name = leah_tables[0].columns[TIME_COLUMN_INDEX]
        
        # Change the "Finals" column name to the time column name in Leah's tables
        rename_final_column(leah_tables, time_column_name)
        
        # Get event names from Leah's tables
        events = [get_event_name(event) for event in full_events]
        
        progress_callback("Matching swimmers between files...", "yellow")

        # For each swimmer in Leah's version, find the corresponding time in Sammy's version
        matched_events = match_swimmers(
            qualifiers_table, 
            leah_tables, 
            events, 
            time_column_name, 
            confirm_callback=confirm_callback,
            progress_callback=progress_callback
        )

        progress_callback("Processing extra swimmers...")

        # Get extra swimmers per event
        extras_per_event = get_extras_per_event(qualifiers_table, events, swimmer_info, matched_events)
        
        # Insert extras into each Leah table before combining
        leah_tables_with_extras = add_extras_to_leah_tables(leah_tables, extras_per_event)

        progress_callback("Generating output file...")

        # Combine all tables into a single output table
        output_table = combine_tables(leah_tables_with_extras, time_column_name)
        
        # Change back the time column name to "Finals" for finals only using the output table
        restore_final_column(output_table)
        
        # Save the output table to an Excel file
        save_output_table_to_excel(output_table, output_path, time_column_name)

        progress_callback(f"✅ FILES PROCESSED SUCCESSFULLY! Output saved as '{output_path}'", "green")

    except Exception as e:
        error_callback(f"❌ ERROR: {str(e)}", "red")
