from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import math
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from leahify_qualifiers import TIME_COLUMN_INDEX, get_leah_tables


QUALIFIER_SLOTS = 6
RESERVE_SLOTS = 2
HEAT_PREFIX = "heat"


@dataclass
class NormalisedSwimmer:
    full_name: str
    age: Any
    seed_time: str
    time: str
    parsed_time: float | None


@dataclass
class EventRanking:
    event_name: str
    rows: list[NormalisedSwimmer]
    qualifier_count: int
    reserve_count: int
    sixth_place_tie_time: float | None


def _as_clean_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _normalise_name(value: Any) -> str:
    text = _as_clean_string(value)
    if not text:
        return ""
    if "," in text:
        surname, first_names = text.split(",", 1)
        return f"{surname.strip()}, {first_names.strip()}"
    return text


def _build_extra_name(first_name: Any, surname: Any) -> str:
    first = _as_clean_string(first_name)
    last = _as_clean_string(surname)
    if not first and not last:
        return ""
    return f"{last}, {first}".strip(", ")


def _parse_dob(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _as_clean_string(value)
    if not text:
        return None

    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None

    return parsed.date()


def _age_from_dob(value: Any) -> str:
    dob = _parse_dob(value)
    if dob is None:
        return ""

    today = date.today()
    age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    return str(age)


def _parse_time_to_seconds(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = _as_clean_string(value)
    if not text:
        return None

    compact = re.sub(r"\s+", "", text)

    if re.fullmatch(r"\d+(\.\d+)?", compact):
        return float(compact)

    if re.fullmatch(r"\d+(,\d+)?", compact):
        return float(compact.replace(",", "."))

    parts = re.split(r"[\.,:;]", compact)
    if any(part == "" for part in parts):
        return None

    if len(parts) == 3:
        try:
            minutes = int(parts[0])
            seconds = int(parts[1])
            hundredths = int(parts[2])
            return (minutes * 60) + seconds + (hundredths / (10 ** len(parts[2])))
        except ValueError:
            return None

    if len(parts) == 2:
        separator_match = re.search(r"[\.,:;]", compact)
        if separator_match and separator_match.group(0) in [":", ";"]:
            try:
                minutes = int(parts[0])
                seconds = float(parts[1])
                return (minutes * 60) + seconds
            except ValueError:
                return None

        # For dot/comma two-part strings (e.g., 10.20), parse as decimal seconds.
        try:
            return float(parts[0] + "." + parts[1])
        except ValueError:
            return None

    return None


def _is_heat_row(row: pd.Series) -> bool:
    first_value = _as_clean_string(row.iloc[0]).lower()
    return first_value.startswith(HEAT_PREFIX)


def _normalise_event_rows(leah_table: pd.DataFrame) -> list[NormalisedSwimmer]:
    normalized_rows: list[NormalisedSwimmer] = []
    in_extras = False

    for _, row in leah_table.iterrows():
        if _is_heat_row(row):
            continue

        first_col = _as_clean_string(row.iloc[0])
        if not first_col:
            continue

        if first_col.lower() == "extra":
            in_extras = True
            continue

        if in_extras:
            full_name = _build_extra_name(row.iloc[0], row.iloc[1])
            age = _age_from_dob(row.iloc[3])
            seed_time = ""
            time = _as_clean_string(row.iloc[TIME_COLUMN_INDEX])
        else:
            full_name = _normalise_name(row.iloc[1])
            age = _as_clean_string(row.iloc[3])
            seed_time = _as_clean_string(row.iloc[4])
            time = _as_clean_string(row.iloc[TIME_COLUMN_INDEX])

        if not full_name:
            continue

        parsed_time = _parse_time_to_seconds(time)
        normalized_rows.append(
            NormalisedSwimmer(
                full_name=full_name,
                age=age,
                seed_time=seed_time,
                time=time,
                parsed_time=parsed_time,
            )
        )

    timed_rows = [row for row in normalized_rows if row.parsed_time is not None]
    non_timed_rows = [row for row in normalized_rows if row.parsed_time is None]

    timed_rows.sort(key=lambda row: row.parsed_time)
    return timed_rows + non_timed_rows


def _expand_with_ties(
    rows: list[NormalisedSwimmer],
    base_count: int,
) -> int:
    if base_count <= 0:
        return 0
    if base_count >= len(rows):
        return len(rows)

    cutoff_time = rows[base_count - 1].parsed_time
    count = base_count

    while count < len(rows) and rows[count].parsed_time == cutoff_time:
        count += 1

    return count


def _build_event_rankings(
    leah_tables: list[pd.DataFrame],
    event_names: list[str],
    progress_callback,
) -> list[EventRanking]:
    rankings: list[EventRanking] = []

    for idx, leah_table in enumerate(leah_tables):
        event_name = event_names[idx]
        rows = _normalise_event_rows(leah_table)

        timed_rows = [row for row in rows if row.parsed_time is not None]

        qualifier_base = min(QUALIFIER_SLOTS, len(timed_rows))
        qualifier_count = _expand_with_ties(timed_rows, qualifier_base)
        sixth_place_tie_time = None

        if qualifier_base == QUALIFIER_SLOTS and qualifier_count > qualifier_base:
            sixth_place_tie_time = timed_rows[QUALIFIER_SLOTS - 1].parsed_time

        if qualifier_count > qualifier_base:
            progress_callback(
                f"WARNING: Tie for qualifiers in '{event_name}'. Including {qualifier_count - qualifier_base} additional swimmer(s).",
                "yellow",
            )

        reserve_start = qualifier_count
        tie_overflow = max(qualifier_count - QUALIFIER_SLOTS, 0)
        effective_reserve_slots = max(RESERVE_SLOTS - tie_overflow, 0)
        reserve_base = min(effective_reserve_slots, max(len(timed_rows) - reserve_start, 0))
        reserve_count = _expand_with_ties(timed_rows[reserve_start:], reserve_base)

        if reserve_count > reserve_base:
            progress_callback(
                f"WARNING: Tie for reserves in '{event_name}'. Including {reserve_count - reserve_base} additional swimmer(s).",
                "yellow",
            )

        rankings.append(
            EventRanking(
                event_name=event_name,
                rows=rows,
                qualifier_count=qualifier_count,
                reserve_count=reserve_count,
                sixth_place_tie_time=sixth_place_tie_time,
            )
        )

    return rankings


def _extract_event_lower_age(event_name: str) -> int | None:
    lower_event = event_name.lower()

    under_match = re.search(r"(\d{1,2})\s*&\s*under", lower_event)
    if under_match:
        return 0

    over_match = re.search(r"(\d{1,2})\s*&\s*over", lower_event)
    if over_match:
        return int(over_match.group(1))

    range_match = re.search(r"(\d{1,2})\s*[-/]\s*(\d{1,2})", lower_event)
    if range_match:
        return int(range_match.group(1))

    return None


def _extract_event_gender_rank(event_name: str) -> int:
    lower_event = event_name.lower()
    if "girls" in lower_event:
        return 0
    if "boys" in lower_event:
        return 1
    return 2


def _sort_rankings_by_age(rankings: list[EventRanking]) -> list[EventRanking]:
    indexed = list(enumerate(rankings))

    def sort_key(item: tuple[int, EventRanking]) -> tuple[int, int, int, int]:
        original_idx, ranking = item
        lower_age = _extract_event_lower_age(ranking.event_name)
        gender_rank = _extract_event_gender_rank(ranking.event_name)
        if lower_age is None:
            return (1, 0, gender_rank, original_idx)
        return (0, lower_age, gender_rank, original_idx)

    sorted_indexed = sorted(indexed, key=sort_key)
    return [ranking for _, ranking in sorted_indexed]


def _save_rankings_to_excel(rankings: list[EventRanking], output_path: str) -> None:
    rankings = _sort_rankings_by_age(rankings)

    rows = []

    for ranking in rankings:
        rows.append([ranking.event_name, "", "", ""])
        rows.append(["Name", "Age", "Seed Time", "Time"])
        for swimmer in ranking.rows:
            rows.append([swimmer.full_name, swimmer.age, swimmer.seed_time, swimmer.time])
        rows.append(["", "", "", ""])

    output_df = pd.DataFrame(rows, columns=["Name", "Age", "Seed Time", "Time"])

    wb = Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    blue_fill = PatternFill(start_color="9FC5E8", end_color="9FC5E8", fill_type="solid")

    for row in dataframe_to_rows(output_df, index=False, header=False):
        ws.append(row)

    event_row_index = 1
    for ranking in rankings:
        header_row_index = event_row_index + 1
        first_data_row = header_row_index + 1

        for offset in range(ranking.qualifier_count):
            row_idx = first_data_row + offset
            for cell in ws[row_idx]:
                cell.fill = green_fill

        if ranking.sixth_place_tie_time is not None:
            for offset in range(ranking.qualifier_count):
                swimmer = ranking.rows[offset]
                if swimmer.parsed_time == ranking.sixth_place_tie_time:
                    row_idx = first_data_row + offset
                    for cell in ws[row_idx]:
                        cell.fill = blue_fill

        reserve_start = first_data_row + ranking.qualifier_count
        for offset in range(ranking.reserve_count):
            row_idx = reserve_start + offset
            for cell in ws[row_idx]:
                cell.fill = yellow_fill

        event_row_index = first_data_row + len(ranking.rows) + 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    wb.save(output_path)


def generate_rankings(
    input_path: str,
    output_path: str,
    progress_callback,
    error_callback,
) -> None:
    try:
        output_path = output_path or "qualifiers_rankings.xlsx"

        progress_callback("Loading Leahify qualifiers output...")
        leah_tables, event_names, _ = get_leah_tables(input_path, None)

        rankings = _build_event_rankings(leah_tables, event_names, progress_callback)

        progress_callback("Generating rankings workbook...")
        _save_rankings_to_excel(rankings, output_path)

        progress_callback(
            f"SUCCESS: Rankings generated. Output saved as '{output_path}'",
            "green",
        )
    except Exception as exc:
        error_callback(f"ERROR: {str(exc)}", "red")
