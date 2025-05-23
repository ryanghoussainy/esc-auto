'''
This file's purpose is to turn Sammy's version of qualifiers into Leah's version,
hence the name leahify.
First, read leah's version, then for each swimmer in that sheet, find the corresponding
time in Sammy's version and add it to the dictionary.
'''

from extract_tables import extract_tables, print_tables, concat_tables, print_first_rows
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re


GROUPS = ["Dolphins", "Herrings", "Seals", "Piranhas", "Marlins"]

QUAL_TABLE_ID = "First name"  # Identifier for the qualifiers table
LEAH_TABLE_ID = "Lane"        # Identifier for the normal tables


def get_qualifiers_table(file: str, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    '''
    Extract the tables from Sammy's version of the qualifiers.
    '''
    qualifiers_table, _, s_info = extract_tables(file, sheet_name, [
        (QUAL_TABLE_ID, 0)
    ])
    return concat_tables(qualifiers_table), s_info


def get_leah_tables(file: str, sheet_name: str) -> tuple[list[pd.DataFrame], list[str]]:
    '''
    Extract the tables from Leah's version of the qualifiers.
    
    Note: the length of the events will be equal to the length
    of the normal tables, so does not include the extra tables.
    '''
    return extract_tables(file, sheet_name, [
        (LEAH_TABLE_ID, 0)
    ], get_events=True)


def leahify_qualifiers(
        sfile: str,
        lfile: str,
) -> list[dict]:
    '''
    Turn Sammy's version of qualifiers into Leah's version.
    '''

    def get_event_name(row_str) -> str:
        '''
        Extract the event name from the input string.
        e.g. "Event  21   Girls 8 & Under 25 SC Meter Breaststroke" -> "25m Breast"
        '''
        pattern = r"\b(25|50|100|200)\s*SC\s*Meter\s*(Freestyle|Backstroke|Breaststroke|Butterfly|IM)"
        match = re.search(pattern, row_str, re.IGNORECASE)

        if match:
            distance = match.group(1)
            stroke = match.group(2)

            stroke_map = {
                "Freestyle": "Free",
                "Backstroke": "Back",
                "Breaststroke": "Breast",
                "Butterfly": "Fly",
                "IM": "IM"
            }
            formatted_stroke = stroke_map[stroke]
            return f"{distance}m {formatted_stroke}"
        else:
            raise ValueError(f"Invalid event name: {row_str}")
        

    def parse_name(name: str) -> tuple[str, str]:
        '''
        Parse the name into first name and surname.
        '''
        names = str(name).split(",")

        if len(names) == 2:
            surname, first_name = names
            
            # Format the names
            first_name = first_name.strip().lower()
            surname = surname.strip().lower()

            return first_name, surname
        else:
            raise ValueError(f"Invalid name: {name}")
        
    # Keep "global" lists for genders
    swimmer_info = {}

    # Get all tables from Sammy's version from each sheet and concatenate them
    qualifiers_tables = []
    for group in GROUPS:
        qualifiers_table, s_info = get_qualifiers_table(sfile, group)
        qualifiers_tables.append(qualifiers_table)

        # Add the swimmer info to the global dictionary
        swimmer_info.update(s_info)
    qualifiers_table = pd.concat(qualifiers_tables, ignore_index=True)

    # Extract tables from Leah's version
    leah_tables, full_events, _ = get_leah_tables(lfile, None)
    events = [get_event_name(event) for event in full_events]

    # Keep track of number of successful matches
    num_matches = 0
    total = 0

    # Keep track of manual and automatic matches
    manual_matches = {}
    automatic_matches = {}

    # Keep track of which events have been matched.
    # This is for the extras table (i.e. which swimmers swam but did not sign up)
    matched_events = {}  # Map from swimmer name (sammy's version) to a list of events

    # For each swimmer in Leah's version, find the corresponding time in Sammy's version
    for tableIdx in range(len(leah_tables)):
        # Get the event
        full_event = full_events[tableIdx]
        event = events[tableIdx]

        ltable = leah_tables[tableIdx]

        for lrowIdx, lrow in ltable.iterrows():
            # Skip NAN rows
            if pd.isnull(lrow["Name"]):
                continue

            # Skip Finals
            if "Time" not in lrow:
                continue

            total += 1

            # Parse the name
            lfirst_names, lsurname = parse_name(lrow["Name"])
            lfirst_name = lfirst_names.split(" ")[0]

            # Check for automatic matches
            if (lfirst_name, lsurname) in automatic_matches:
                sfirst_name, ssurname = automatic_matches[(lfirst_name, lsurname)]
                swimmer = qualifiers_table[
                    (qualifiers_table["First name"] == sfirst_name) & 
                    (qualifiers_table["Surname"] == ssurname)
                ]
            else:
                # Get all close matches with fuzzy matching
                scores = []
                for _, srow in qualifiers_table.iterrows():
                    # Skip automatic and manual matches
                    if (srow["First name"], srow["Surname"]) in automatic_matches.values() or \
                        (srow["First name"], srow["Surname"]) in manual_matches.values():
                        continue
                    
                    scores.append((srow["First name"], srow["Surname"], fuzz.ratio(lfirst_name + " " + lsurname, srow["First name"].lower() + " " + srow["Surname"].lower())))
                scores.sort(key=lambda x: x[2], reverse=True)

                # If the first one isn't identical, then ask the user
                if scores[0][2] < 100:
                    # Check for manual matches
                    if (lfirst_name, lsurname) in manual_matches:
                        sfirst_name, ssurname = manual_matches[(lfirst_name, lsurname)]
                        swimmer = qualifiers_table[
                            (qualifiers_table["First name"] == sfirst_name) &
                            (qualifiers_table["Surname"] == ssurname)
                        ]
                    else:
                        print("\033[93m" +
                            f"Trying to match... {lfirst_name.capitalize()} {lsurname.capitalize()}" +
                            "\033[0m")
                        
                        # Keep asking the user to say yes or no to each match
                        for _, (sfirst_name, ssurname, score) in enumerate(scores):
                            # Ask the user if the match is correct
                            print(f"Is this the right match? \033[93m{lfirst_name.capitalize()} {lsurname.capitalize()}\033[0m -> \033[93m{sfirst_name.capitalize()} {ssurname.capitalize()}\033[0m (similarity score: {score}%)")
                            
                            # Get user input
                            match = input("(y/n/exit): ")
                            while match.lower() not in ["y", "n", "exit"]:
                                print("Invalid input")
                                match = input("(y/n/exit): ")

                            if match.lower() == "y":
                                swimmer = qualifiers_table[
                                    (qualifiers_table["First name"] == sfirst_name) &
                                    (qualifiers_table["Surname"] == ssurname)
                                ]

                                # Add to manual matches
                                manual_matches[(lfirst_name, lsurname)] = (sfirst_name, ssurname)

                                break
                            elif match.lower() == "exit":
                                # Exit the program
                                exit()

                        if swimmer.empty:
                            raise ValueError(f"No swimmer found: {lfirst_name.capitalize()} {lsurname.capitalize()}")
                    
                else:
                    swimmer = qualifiers_table[
                        (qualifiers_table["First name"] == scores[0][0]) &
                        (qualifiers_table["Surname"] == scores[0][1])
                    ]

                    # Add to automatic matches
                    automatic_matches[(lfirst_name, lsurname)] = (scores[0][0], scores[0][1])
            
            # Increment number of matches
            num_matches += 1

            # Add the event to the matched events
            sfirst_name = swimmer["First name"].values[0]
            ssurname = swimmer["Surname"].values[0]
            if (sfirst_name, ssurname) in matched_events:
                matched_events[(sfirst_name, ssurname)].append(event)
            else:
                matched_events[(sfirst_name, ssurname)] = [event]

            # Print in green
            print("\033[92m" +
                  f"Successfully matched: {lfirst_name.capitalize()} {lsurname.capitalize()}" +
                  "\033[0m")
            
            # Get the time
            time = swimmer[event].values[0]

            # Set the column type to string
            leah_tables[tableIdx] = leah_tables[tableIdx].astype({"Time": str})
            
            # Set the time if it's not nan
            if not pd.isnull(time):
                leah_tables[tableIdx].at[lrowIdx, "Time"] = time
            else:
                leah_tables[tableIdx].at[lrowIdx, "Time"] = "DNS"

    print(f"Number of matches: {num_matches} / {total}")

    # Concatenate all tables into a single table, adding a header row for each one
    ltable = leah_tables[0]
    output_table = pd.DataFrame(columns=ltable.columns)

    header_row = pd.DataFrame([ltable.columns], columns=ltable.columns)

    for tableIdx in range(len(leah_tables)):
        # Add the header row
        ltable = leah_tables[tableIdx]
        ltable_with_header = pd.concat([ltable.iloc[:1], header_row, ltable.iloc[1:]])
        output_table = pd.concat([output_table, ltable_with_header])

    # Get rid of nan values in the Time column
    output_table["Time"] = output_table["Time"].replace("nan", "")

    # Keep track of extra swimmers for each event
    extras_per_event = {} # Map from (event name, age range, gender) to a list of swimmers

    for idx, srow in qualifiers_table.iterrows():
        # Get all events swam by the swimmer
        events_swam = set(event for event in set(events) if not pd.isnull(srow[event]) and srow[event] != "DNS")

        # Remove the 200m Free since we only care about the normal events
        if "200m Free" in events_swam:
            events_swam.remove("200m Free")

        # Missing events
        missing_events = []

        # Check if they have been matched for the same events
        for ev in events_swam:
            if (srow["First name"], srow["Surname"]) not in matched_events or ev not in matched_events[(srow["First name"], srow["Surname"])]:
                missing_events.append(ev)

        # Add the swimmer to the extras per event
        for missing_event in missing_events:
            # Get swimmer age range and gender
            age_from, age_to, gender = swimmer_info[(srow["First name"], srow["Surname"])]
            # Add the swimmer to the extras per event table
            extras_per_event.setdefault((missing_event, age_from, age_to, gender), []).append(srow)

    # Instead of modifying output_table in-place, build a new list of rows
    new_rows = []
    prv_event = prv_age_from = prv_age_to = prv_gender = None

    for idx, row in output_table.iterrows():
        # Insert extras for the previous event before the next event header
        if row.iloc[0].startswith("Event"):
            if prv_event is not None:
                key = (prv_event, prv_age_from, prv_age_to, prv_gender)
                if key in extras_per_event:
                    # Insert a label row
                    new_rows.append(["EXTRA"] + [""] * (output_table.shape[1] - 1))
                    # Insert each extra swimmer row (convert Series to list)
                    for extra_row in extras_per_event[key]:
                        new_rows.append(extra_row[["First name", "Surname", "ASA", "DOB", "Group", prv_event]].tolist() + [""] * (output_table.shape[1] - 6))

            # Extract the event name
            prv_event = get_event_name(row.iloc[0])
            # Extract the age range
            age_range = re.search(r"\b\d{1,2}\s*&\s*(Under|Over|under|over)|\b\d{1,2}\s*-\s*\d{1,2}", row.iloc[0])
            if age_range:
                age_range = age_range.group(0)
                if "under" in age_range.lower():
                    prv_age_from = 0
                    prv_age_to = int(age_range.split("&")[0].strip())
                elif "over" in age_range.lower():
                    prv_age_from = int(age_range.split("&")[0].strip())
                    prv_age_to = 99
                else:
                    if "-" in age_range:
                        prv_age_from, prv_age_to = map(int, age_range.split("-"))
                    elif "/" in age_range:
                        prv_age_from, prv_age_to = map(int, age_range.split("/"))
                    else:
                        raise ValueError(f"Could not extract age range. Did not find - or /: {row.iloc[0]}")
            elif "200" in row.iloc[0]:
                pass
            else:
                raise ValueError(f"Could not extract age range from event name: {row.iloc[0]}")

            prv_gender = "boys" if "boys" in row.iloc[0].lower() else "girls"

        # Always add the current row
        new_rows.append(list(row))

    # Rebuild the output_table from new_rows
    output_table = pd.DataFrame(new_rows, columns=output_table.columns)

    # Add a border to the table
    wb = Workbook()
    ws = wb.active
    big_font = Font(size=18)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in dataframe_to_rows(output_table, index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=1, max_row=len(output_table), min_col=1, max_col=len(output_table.columns)):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if cell.column == 6 and cell.value != "Time":
                cell.font = big_font
            if cell.row > 1 and cell.column == 1 and cell.value == "EXTRA":
                cell.fill = yellow_fill

    # Save the output table
    wb.save("output.xlsx")


if __name__ == '__main__':

    leahify_qualifiers('examples/1_doc.xlsx', 'examples/2_docLeah.xls')

