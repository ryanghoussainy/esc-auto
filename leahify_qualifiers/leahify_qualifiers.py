'''
This file's purpose is to turn Sammy's version of qualifiers into Leah's version,
hence the name leahify.
First, read leah's version, then for each swimmer in that sheet, find the corresponding
time in Sammy's version and add it to the dictionary.
'''

from extract_tables import extract_tables, print_tables, concat_tables, print_first_rows
from printing import print_colour, RED, YELLOW, GREEN
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re


GROUPS = ["Dolphins", "Herrings", "Seals", "Piranhas", "Marlins"]

QUAL_TABLE_ID = "First name"  # Identifier for the qualifiers table
LEAH_TABLE_ID = "Lane"        # Identifier for the normal tables

VALID_MATCH_INPUTS = ["y", "n", "exit"]  # Valid inputs for matching swimmers
VALID_MATCH_INPUTS_STR = f"({'/'.join(VALID_MATCH_INPUTS)}): "  # e.g. "(y, n, exit)"

REGEX_AGE_RANGE_LEAH = r"\b\d{1,2}\s*&\s*(Under|Over|under|over)|\b\d{1,2}\s*-\s*\d{1,2}" # Regex for age range
REGEX_EVENT_NAME = r"\b(25|50|100|200)\s*SC\s*Meter\s*(Freestyle|Backstroke|Breaststroke|Butterfly|IM)"


def get_event_name(row_str) -> str:
    '''
    Extract the event name from the input string.
    e.g. "Event  21   Girls 8 & Under 25 SC Meter Breaststroke" -> "25m Breast"
    '''
    match = re.search(REGEX_EVENT_NAME, row_str, re.IGNORECASE)

    if match:
        distance = match.group(1)
        stroke = match.group(2)

        stroke_map = {
            "Freestyle": "Free",
            "Backstroke": "Back",
            "Breaststroke": "Breast",
            "Butterfly": "Fly",
            "IM": "IM"
        }
        formatted_stroke = stroke_map[stroke]
        return f"{distance}m {formatted_stroke}"
    else:
        raise ValueError(f"Invalid event name: {row_str}")    

def parse_name(name: str) -> tuple[str, str]:
    '''
    Parse the name into first name and surname.
    '''
    names = str(name).split(",")

    if len(names) == 2:
        surname, first_name = names
        
        # Format the names
        first_name = first_name.strip().lower()
        surname = surname.strip().lower()

        return first_name, surname
    else:
        raise ValueError(f"Invalid name: {name}")

def get_qualifiers_table(file: str, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    '''
    Extract the tables from Sammy's version of the qualifiers.
    '''
    qualifiers_table, _, s_info = extract_tables(file, sheet_name, [
        (QUAL_TABLE_ID, 0)
    ])
    return concat_tables(qualifiers_table), s_info

def get_leah_tables(file: str, sheet_name: str) -> tuple[list[pd.DataFrame], list[str]]:
    '''
    Extract the tables from Leah's version of the qualifiers.
    
    Note: the length of the events will be equal to the length
    of the normal tables, so does not include the extra tables.
    '''
    return extract_tables(file, sheet_name, [
        (LEAH_TABLE_ID, 0)
    ], get_events=True)

def load_qualifiers(sfile: str) -> tuple[pd.DataFrame, dict]:
    '''
    Load all qualifiers tables and swimmer info (age from, age to, gender) from the Excel file.
    '''
    qualifiers_tables = []
    swimmer_info = {}
    for group in GROUPS:
        qualifiers_table, s_info = get_qualifiers_table(sfile, group)
        qualifiers_tables.append(qualifiers_table)
        swimmer_info.update(s_info)
    # Concatenate all tables into a single table
    qualifiers_tables = pd.concat(qualifiers_tables, ignore_index=True)

    return qualifiers_tables, swimmer_info

def get_close_matches(
        qualifiers_table: pd.DataFrame,
        lfirst_name: str,
        lsurname: str,
        automatic_matches: dict,
        manual_matches: dict,
) -> list[tuple[str, str, int]]:
    '''
    Get the closest matches for a swimmer in Sammy's version.
    '''
    scores = []
    for _, srow in qualifiers_table.iterrows():
        # Skip automatic and manual matches
        if (srow["First name"], srow["Surname"]) in automatic_matches.values() or \
            (srow["First name"], srow["Surname"]) in manual_matches.values():
            continue
        
        scores.append((srow["First name"], srow["Surname"], fuzz.ratio(lfirst_name + " " + lsurname, srow["First name"].lower() + " " + srow["Surname"].lower())))
    scores.sort(key=lambda x: x[2], reverse=True)
    return scores

def match_swimmers(
        qualifiers_table: pd.DataFrame,
        leah_tables: list[pd.DataFrame],
        events: list[str],
) -> dict:
    '''
    Match swimmers from Leah's version to Sammy's version.
    '''
    # For each swimmer in Leah's version, find the corresponding time in Sammy's version
    # Keep track of number of successful matches
    num_matches = 0
    total = 0

    # Keep track of manual and automatic matches
    manual_matches = {} # manual matches are from user input
    automatic_matches = {} # automatic matches are from exact matches

    # Keep track of which events have been matched.
    # This is for the extras table (i.e. which swimmers swam but did not sign up)
    matched_events = {}  # Map from swimmer name (sammy's version) to a list of events

    # Each table corresponds to an event
    # Iterate over each table and each row to get each swimmer
    for tableIdx in range(len(leah_tables)):
        # Get the event
        event = events[tableIdx]

        ltable = leah_tables[tableIdx]

        # Iterate over each row in the table (i.e. each swimmer)
        for lrowIdx, lrow in ltable.iterrows():
            # Skip NAN rows
            if pd.isnull(lrow["Name"]):
                continue

            # Skip Finals (i.e. the 200m Free one-off final)
            if "Time" not in lrow:
                continue

            total += 1

            # Parse the swimmer's name (Leah's version)
            lfirst_names, lsurname = parse_name(lrow["Name"])
            # Ignore middle names
            lfirst_name = lfirst_names.split(" ")[0]

            # Check for automatic matches
            if (lfirst_name, lsurname) in automatic_matches:
                sfirst_name, ssurname = automatic_matches[(lfirst_name, lsurname)]
                swimmer = qualifiers_table[
                    (qualifiers_table["First name"] == sfirst_name) & 
                    (qualifiers_table["Surname"] == ssurname)
                ]
            else:
                # Get all close matches
                scores = get_close_matches(qualifiers_table, lfirst_name, lsurname, automatic_matches, manual_matches)

                # If the first one is identical, then use it
                if scores[0][2] == 100:
                    swimmer = qualifiers_table[
                        (qualifiers_table["First name"] == scores[0][0]) &
                        (qualifiers_table["Surname"] == scores[0][1])
                    ]

                    # Add to automatic matches
                    automatic_matches[(lfirst_name, lsurname)] = (scores[0][0], scores[0][1])
                    
                # If it's already been manually matched, then use that
                elif (lfirst_name, lsurname) in manual_matches:
                    sfirst_name, ssurname = manual_matches[(lfirst_name, lsurname)]
                    swimmer = qualifiers_table[
                        (qualifiers_table["First name"] == sfirst_name) &
                        (qualifiers_table["Surname"] == ssurname)
                    ]

                # Otherwise, prompt the user to match manually
                else:
                    print_colour(YELLOW, f"Trying to match... {lfirst_name.capitalize()} {lsurname.capitalize()}")
                    
                    # Keep asking the user to say `yes`, `no` or `exit` to each match
                    for _, (sfirst_name, ssurname, score) in enumerate(scores):
                        # Ask the user if the match is correct
                        print("Is this the right match? ", end="")
                        print_colour(YELLOW, f"{lfirst_name.capitalize()} {lsurname.capitalize()}", end="")
                        print_colour(YELLOW, f" -> {sfirst_name.capitalize()} {ssurname.capitalize()}", end="")
                        print(f" (similarity score: {score}%)")
                        match = input(VALID_MATCH_INPUTS_STR)

                        # Check if the input is valid
                        while match.lower() not in VALID_MATCH_INPUTS:
                            print("Invalid input")
                            match = input(VALID_MATCH_INPUTS_STR)

                        # Handle exiting
                        if  match.lower() == "exit":
                            # Exit the program
                            exit()

                        # If the user confirms a match, then use it.
                        # Otherwise, continue to the next match
                        elif match.lower() == "y":
                            swimmer = qualifiers_table[
                                (qualifiers_table["First name"] == sfirst_name) &
                                (qualifiers_table["Surname"] == ssurname)
                            ]

                            # Add to manual matches
                            manual_matches[(lfirst_name, lsurname)] = (sfirst_name, ssurname)

                            break

                    if swimmer.empty:
                        raise ValueError(f"No swimmer found: {lfirst_name.capitalize()} {lsurname.capitalize()}")
            
            # Increment number of matches
            num_matches += 1

            # Add the event to the matched events
            sfirst_name = swimmer["First name"].values[0]
            ssurname = swimmer["Surname"].values[0]
            if (sfirst_name, ssurname) in matched_events:
                matched_events[(sfirst_name, ssurname)].append(event)
            else:
                matched_events[(sfirst_name, ssurname)] = [event]

            # Print in green
            print_colour(GREEN, f"Successfully matched: {lfirst_name.capitalize()} {lsurname.capitalize()}")
            
            # Get the time
            time = swimmer[event].values[0]

            # Set the column type to string
            leah_tables[tableIdx] = leah_tables[tableIdx].astype({"Time": str})
            
            # Set the time if it's not nan
            if not pd.isnull(time):
                leah_tables[tableIdx].at[lrowIdx, "Time"] = time
            else:
                leah_tables[tableIdx].at[lrowIdx, "Time"] = "DNS"

    # Print the number of matches
    print(f"Number of matches: {num_matches} / {total}")

    return matched_events

def combine_tables(leah_tables: list[pd.DataFrame]) -> pd.DataFrame:
    '''
    Concatenate all tables into a single table.
    '''
    ltable = leah_tables[0]
    output_table = pd.DataFrame(columns=ltable.columns)

    header_row = pd.DataFrame([ltable.columns], columns=ltable.columns)

    for tableIdx in range(len(leah_tables)):
        # Add the header row
        ltable = leah_tables[tableIdx]
        ltable_with_header = pd.concat([ltable.iloc[:1], header_row, ltable.iloc[1:]])
        output_table = pd.concat([output_table, ltable_with_header])

    # Get rid of nan values in the Time column
    output_table["Time"] = output_table["Time"].replace("nan", "")

    return output_table

def get_extras_per_event(
        qualifiers_table: pd.DataFrame,
        events: list[str],
        swimmer_info: dict,
        matched_events: dict,        
) -> dict:
    # Keep track of extra swimmers for each event
    extras_per_event = {} # Map from (event name, age range, gender) to a list of swimmers

    for idx, srow in qualifiers_table.iterrows():
        # Get all events swam by the swimmer
        events_swam = set(event for event in set(events) if not pd.isnull(srow[event]) and srow[event] != "DNS")

        # Remove the 200m Free since we only care about the normal events
        if "200m Free" in events_swam:
            events_swam.remove("200m Free")

        # Missing events
        missing_events = []

        # Check if they have been matched for the same events
        for ev in events_swam:
            if (srow["First name"], srow["Surname"]) not in matched_events or ev not in matched_events[(srow["First name"], srow["Surname"])]:
                missing_events.append(ev)

        # Add the swimmer to the extras per event
        for missing_event in missing_events:
            # Get swimmer age range and gender
            age_from, age_to, gender = swimmer_info[(srow["First name"], srow["Surname"])]
            # Add the swimmer to the extras per event table
            extras_per_event.setdefault((missing_event, age_from, age_to, gender), []).append(srow)

    return extras_per_event

def add_extras_to_leah_tables(
    leah_tables: list[pd.DataFrame],
    extras_per_event: dict,
) -> list[pd.DataFrame]:
    """
    Insert extras into each Leah table before combining.
    """
    
    for i in range(len(leah_tables)):
        leah_table = leah_tables[i]
        event_cell = leah_table.iloc[0, 0]

        # Extract event name
        event_name = get_event_name(event_cell)

        # Extract age range
        age_range = re.search(REGEX_AGE_RANGE_LEAH, event_cell)
        if age_range:
            age_range = age_range.group(0)
            if "under" in age_range.lower():
                age_from = 0
                age_to = int(age_range.split("&")[0].strip())
            elif "over" in age_range.lower():
                age_from = int(age_range.split("&")[0].strip())
                age_to = 99
            else:
                if "-" in age_range:
                    age_from, age_to = map(int, age_range.split("-"))
                elif "/" in age_range:
                    age_from, age_to = map(int, age_range.split("/"))
                else:
                    raise ValueError(f"Could not extract age range. Did not find - or /: {event_cell}")
        elif "200" in event_cell:
            pass
        else:
            raise ValueError(f"Could not extract age range from event name: {event_cell}")

        gender = "boys" if "boys" in event_cell.lower() else "girls"
        key = (event_name, age_from, age_to, gender)

        # If extras_per_event has an entry, add extras to the table
        if key in extras_per_event:
            # Add extra label row
            extra_label_row = pd.DataFrame([["EXTRA"] + [""] * (len(leah_table.columns) - 1)], columns=leah_table.columns)
            leah_tables[i] = pd.concat([leah_tables[i], extra_label_row], ignore_index=True)

            # For each extra swimmer, create a row
            for extra_row in extras_per_event[key]:
                # Add the correct columns to insert into Leah table
                extra_row = extra_row[["First name", "Surname", "ASA", "DOB", "Group", event_name]]
                
                # Pad the row with empty strings to match the number of columns in leah_table
                padded_row = list(extra_row.values) + [""] * (len(leah_table.columns) - len(extra_row.values))
                extra_row_df = pd.DataFrame([padded_row], columns=leah_table.columns)

                # Add it to the Leah table
                leah_tables[i] = pd.concat([leah_tables[i], extra_row_df], ignore_index=True)

    return leah_tables

def save_output_table_to_excel(
        output_table: pd.DataFrame,
        filename: str,
) -> None:
    '''
    Save the output table to an Excel file.
    '''
    # Add a border to the table
    wb = Workbook()
    ws = wb.active
    big_font = Font(size=18)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r in dataframe_to_rows(output_table, index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=1, max_row=len(output_table), min_col=1, max_col=len(output_table.columns)):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if cell.column == 6 and cell.value != "Time":
                cell.font = big_font
            if cell.row > 1 and cell.column == 1 and cell.value == "EXTRA":
                cell.fill = yellow_fill

    # Save the output table
    wb.save(filename)

def leahify_qualifiers(
        sfile: str,
        lfile: str,
) -> None:
    '''
    Turn Sammy's version of qualifiers into Leah's version.
    '''
    # Load qualifier times (Sammy's version)
    # swimmer_info is a map from swimmer to (age from, age to, gender)
    qualifiers_table, swimmer_info = load_qualifiers(sfile)

    # Extract tables from Leah's version
    leah_tables, full_events, _ = get_leah_tables(lfile, None)
    events = [get_event_name(event) for event in full_events]

    # For each swimmer in Leah's version, find the corresponding time in Sammy's version
    matched_events = match_swimmers(qualifiers_table, leah_tables, events)

    # Get extra swimmers per event
    extras_per_event = get_extras_per_event(qualifiers_table, events, swimmer_info, matched_events)

    # Insert extras into each Leah table before combining
    leah_tables_with_extras = add_extras_to_leah_tables(leah_tables, extras_per_event)

    # Combine all tables into a single output table
    output_table = combine_tables(leah_tables_with_extras)

    # Save the output table to an Excel file
    save_output_table_to_excel(output_table, "output.xlsx")


if __name__ == '__main__':

    leahify_qualifiers('examples/1_doc.xlsx', 'examples/2_docLeah.xls')

